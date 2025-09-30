import os
import json
import re
from openpyxl import load_workbook, Workbook 
import sys 

# --- Utilitário de Caminho para PyInstaller/Desenvolvimento ---

def _get_resource_path(relative_path):
    """
    Obtém o caminho de recursos SOMENTE LEITURA.
    Busca o arquivo no ambiente PyInstaller (sys._MEIPASS) ou no diretório do script.
    Usado para MODELO.XLSX, LOGO.PNG e arquivos JSON pré-configurados.
    """
    if hasattr(sys, '_MEIPASS'):
        # Ambiente PyInstaller
        return os.path.join(sys._MEIPASS, relative_path)
    
    # Ambiente de Desenvolvimento (Busca na pasta onde backend_data.py está)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(script_dir, relative_path)

def _get_persistence_path(relative_path):
    """
    Obtém o caminho para arquivos de dados de leitura/escrita (persistência).
    Sempre usa o Current Working Directory (CWD) onde o usuário executa o programa.
    Usado para CLIENTES.JSON, ESTADO.JSON e SAIDA_FOLDER.
    """
    return os.path.join(os.getcwd(), relative_path)

# ---------------------------------------------


# --- Constantes de Arquivos ---
CLIENTES_FILE = "clientes.json"
ESTADO_FILE = "estado.json"
MODELO_FILE = "modelo.xlsx" # Nome do arquivo que será empacotado
SAIDA_FOLDER = "Notas_de_Credito_Geradas"
TEMPLATES_FILE = "templates.json" # NOVO ARQUIVO DE TEMPLATES

# --- Dados Iniciais ---
INITIAL_ESTADO = {
    "ultima_fatura": 1,
    "ultima_descricao": "DESCONTO COMERCIAL REFERENTE A ACERTO COMERCIAL DE PRODUTOS.",
}

# Dados iniciais de templates para a primeira execução
INITIAL_TEMPLATES = [
    {"nome": "Desconto Comercial Padrão", "descricao": "DESCONTO COMERCIAL REFERENTE A ACERTO COMERCIAL DE PRODUTOS."},
    {"nome": "Bonificação por Volume", "descricao": "BONIFICAÇÃO POR VOLUME DE COMPRAS CONFORME ACORDO COMERCIAL DO PERÍODO."},
    {"nome": "Ajuste de Preço", "descricao": "AJUSTE NO VALOR DA FATURA DEVIDO À ALTERAÇÃO DE PREÇOS NA DATA DE EMISSÃO."}
]

# --- Gerenciamento de Dados (JSON) ---

def load_clientes():
    """
    Carrega a lista de clientes. Tenta carregar do diretório de PERSISTÊNCIA (CWD).
    Se não existir, verifica se há dados pré-configurados no RECURSO (PyInstaller).
    """
    # 1. Tenta carregar do CWD (dados persistentes criados pelo usuário)
    persistence_path = _get_persistence_path(CLIENTES_FILE)
    if os.path.exists(persistence_path):
        try:
            with open(persistence_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    
    # 2. Se não encontrou dados persistentes, tenta carregar o arquivo EMBUTIDO (clientes pré-configurados)
    resource_path = _get_resource_path(CLIENTES_FILE)
    if os.path.exists(resource_path):
        try:
            with open(resource_path, 'r', encoding='utf-8') as f:
                print(f"AVISO: Clientes carregados do arquivo embutido: {resource_path}")
                return json.load(f)
        except Exception:
            pass

    return []

def save_clientes(clientes):
    """Salva a lista de clientes no arquivo JSON de PERSISTÊNCIA (CWD)."""
    try:
        persistence_path = _get_persistence_path(CLIENTES_FILE)
        with open(persistence_path, 'w', encoding='utf-8') as f:
            json.dump(clientes, f, indent=4, ensure_ascii=False)
    except Exception as e:
        print(f"Erro ao salvar clientes: {e}")

def load_estado():
    """
    Carrega o estado. Tenta carregar do diretório de PERSISTÊNCIA (CWD).
    Se não existir, tenta carregar o arquivo EMBUTIDO (estado pré-configurado).
    """
    # 1. Tenta carregar do CWD (dados persistentes)
    persistence_path = _get_persistence_path(ESTADO_FILE)
    if os.path.exists(persistence_path):
        try:
            with open(persistence_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass

    # 2. Se não encontrou, tenta carregar o arquivo EMBUTIDO (se estiver empacotado)
    resource_path = _get_resource_path(ESTADO_FILE)
    if os.path.exists(resource_path):
        try:
            with open(resource_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
            
    return INITIAL_ESTADO

def save_estado(estado):
    """Salva o estado da última fatura e descrição no arquivo JSON de PERSISTÊNCIA (CWD)."""
    try:
        persistence_path = _get_persistence_path(ESTADO_FILE)
        with open(persistence_path, 'w', encoding='utf-8') as f:
            json.dump(estado, f, indent=4, ensure_ascii=False)
    except Exception as e:
        print(f"Erro ao salvar estado: {e}")


# --- NOVAS FUNÇÕES DE TEMPLATE ---

def load_templates():
    """
    Carrega a lista de templates. Tenta carregar do CWD, se falhar, retorna os templates iniciais.
    """
    # Tenta carregar do CWD (onde os templates criados pelo usuário são salvos)
    persistence_path = _get_persistence_path(TEMPLATES_FILE)
    if os.path.exists(persistence_path):
        try:
            with open(persistence_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass

    # Retorna o modelo inicial se não houver arquivo persistente
    return INITIAL_TEMPLATES

def save_templates(templates):
    """Salva a lista de templates no arquivo JSON de PERSISTÊNCIA (CWD)."""
    try:
        persistence_path = _get_persistence_path(TEMPLATES_FILE)
        with open(persistence_path, 'w', encoding='utf-8') as f:
            json.dump(templates, f, indent=4, ensure_ascii=False)
    except Exception as e:
        print(f"Erro ao salvar templates: {e}")

# --- Criação do Modelo (sem alteração) ---

def _create_initial_model():
    """Cria um arquivo modelo XLSX mínimo se não existir. Apenas em DEV."""
    
    # Se estiver empacotado, NUNCA cria o modelo, pois ele deve estar embutido.
    if hasattr(sys, '_MEIPASS'):
        return
        
    model_path = _get_resource_path(MODELO_FILE)
    
    if not os.path.exists(model_path):
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "MODELO" 
            
            # Células de preenchimento obrigatórias (simplificado)
            ws['H9'] = 'DATA'
            ws['K9'] = 'FATURA'
            ws['A15'] = 'NOME_CLIENTE'
            
            wb.save(model_path)
            print(f"AVISO: Arquivo modelo '{MODELO_FILE}' criado para uso em DEV. SUBSTITUA pelo seu modelo real.")
        except Exception as e:
            print(f"Erro ao criar modelo inicial: {e}")

# --- Processamento de XLSX (sem alteração) ---

def process_and_save_note(data_input, invoice_number, client_code, client_name, description_text, value_float, estado):
    """
    Carrega o modelo (via PyInstaller ou ambiente local), preenche os dados e salva o novo arquivo XLSX.
    """
    
    # Obtém o caminho correto do modelo (PyInstaller ou local)
    model_path = _get_resource_path(MODELO_FILE)
    
    # Se o modelo não existe, tenta criar o fallback (só em DEV)
    if not os.path.exists(model_path):
        _create_initial_model()
        
    # Tenta obter o caminho novamente após o fallback
    model_path = _get_resource_path(MODELO_FILE)

    if not os.path.exists(model_path):
          return False, f"Erro Fatal: O arquivo modelo '{MODELO_FILE}' não foi encontrado em: {model_path}."


    try:
        # DIAGNÓSTICO: Mostra de onde o modelo está sendo lido.
        print(f"\n--- DIAGNÓSTICO: Carregando modelo de: {model_path} ---")

        # Carrega o modelo
        wb = load_workbook(model_path)
        ws = wb.active
        
        # 1. Preparação dos Nomes (para Planilha e Arquivo)
        
        cleaned_name = re.sub(r'[\\/?*\[\]\':]', '', client_name).strip()
        name_parts = [p for p in cleaned_name.split() if p] 
        
        if len(name_parts) >= 2:
            base_name = f"{name_parts[0]}_{name_parts[1]}"
        elif len(name_parts) == 1:
            base_name = name_parts[0]
        else:
            base_name = "CLIENTE_SEM_NOME"


        # 1.2 Renomeia a Planilha (Tab)
        sheet_name_raw = f"{base_name}_{invoice_number}"
        new_sheet_name = sheet_name_raw[:31].replace(' ', '_')
        ws.title = new_sheet_name

        # 2. Preenchimento de Células (Usando os campos que o usuário pediu)
        
        data_map = {
            'H9': data_input,                  # Data
            'K9': invoice_number,              # Número da Fatura (Topo)
            'A13': client_code,                # Código do Cliente
            'J52': client_code,                # Código do Cliente (Parte inferior)
            'A15': client_name,                # Nome/Razão Social (Faturar a)
            'G15': client_name,                # Nome/Razão Social (Enviado a)
            'B28': description_text,           # Descrição/Histórico
            'K50': value_float,                # Valor da Fatura (Número)
            'L52': invoice_number,             # Número da Fatura (Parte inferior)
        }

        for cell, value in data_map.items():
            try:
                ws[cell] = value
            except Exception:
                 pass 

        # Formata o valor como moeda
        try:
            ws['K50'].number_format = 'R$ #,##0.00'
        except:
            pass 

        # 3. Define o Caminho de Saída (CWD para persistência)
        persistence_saida_folder = _get_persistence_path(SAIDA_FOLDER)
        if not os.path.exists(persistence_saida_folder):
            os.makedirs(persistence_saida_folder)

        output_filename = f"{base_name}_{invoice_number}.xlsx"
        output_path = os.path.join(persistence_saida_folder, output_filename)
        
        # 4. Salva o Arquivo
        wb.save(output_path)

        # 5. Atualiza o Estado (Próxima Fatura e Descrição)
        try:
            current_invoice = int(invoice_number)
            estado['ultima_fatura'] = current_invoice + 1
            estado['ultima_descricao'] = description_text
            save_estado(estado) # Salva o estado usando a nova função (CWD)
        except ValueError:
            pass 

        return True, output_path

    except Exception as e:
        return False, f"Erro ao processar o arquivo XLSX: {type(e).__name__}: {e}"
