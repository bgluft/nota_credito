import os
import json
import datetime
import re
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, Toplevel, Listbox, Scrollbar

# Bibliotecas externas necessárias:
# pip install customtkinter openpyxl

try:
    import customtkinter as ctk
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Erro de importação: {e}")
    print("Por favor, instale as bibliotecas necessárias: pip install customtkinter openpyxl")
    exit()

# --- Configurações de Arquivo ---
CLIENTES_FILE = "clientes.json"
ESTADO_FILE = "estado.json"
MODELO_FILE = "modelo.xlsx"
SAIDA_FOLDER = "Notas_de_Credito_Geradas"

# --- Cores para Listbox (Zebrado) ---
COLOR_LIGHT = "#E0E0E0"  # Cinza claro para linhas pares
COLOR_DARK = "#FFFFFF"   # Branco para linhas ímpares

# --- Funções de Backend/Persistência ---

def load_data(filename, default_data):
    """Carrega dados de um arquivo JSON. Se não existir, retorna dados padrão."""
    try:
        if os.path.exists(filename):
            with open(filename, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            return default_data
    except Exception as e:
        print(f"Erro ao carregar {filename}: {e}")
        return default_data

def save_data(filename, data):
    """Salva dados em um arquivo JSON."""
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
    except Exception as e:
        print(f"Erro ao salvar {filename}: {e}")

# Gerenciamento de Clientes
def load_clientes():
    """Carrega a lista de clientes."""
    return load_data(CLIENTES_FILE, [])

def save_clientes(clientes):
    """Salva a lista de clientes."""
    save_data(CLIENTES_FILE, clientes)

# Gerenciamento de Estado (Fatura e Descrição)
def load_estado():
    """Carrega o estado do programa (última fatura e descrição)."""
    return load_data(ESTADO_FILE, {"ultima_fatura": 1, "ultima_descricao": "Devolução de Mercadoria conforme NFe [NUMERO NFE] de [DATA NFE]."})

def save_estado(estado):
    """Salva o estado atual do programa."""
    save_data(ESTADO_FILE, estado)

# --- Criação do Arquivo Modelo XLSX (Se não existir) ---

def create_initial_xlsx_template():
    """Cria um arquivo modelo XLSX mínimo para garantir a funcionalidade."""
    if os.path.exists(MODELO_FILE):
        return

    print(f"Criando {MODELO_FILE} inicial. Por favor, substitua-o pelo seu modelo real.")
    wb = Workbook()
    # Define o título da primeira planilha para algo que não seja o padrão 'Sheet'
    ws = wb.active
    ws.title = "Modelo_Base"

    # Define as células que o script irá preencher
    # H9: Data
    ws['H9'] = 'Data'
    # K9: Número da Fatura (Topo)
    ws['K9'] = 'Fatura N° (Topo)'
    # A13: Código do Cliente (Topo)
    ws['A13'] = 'Código Cliente (Topo)'
    # A15 e G15: Nome do Cliente (Faturar a: e Enviado a:)
    ws['A15'] = 'Nome Cliente (Razão Social - Faturar)'
    ws['G15'] = 'Nome Cliente (Razão Social - Enviado)' 
    # B28: Descrição/Histórico
    ws['B28'] = 'Descrição/Histórico'
    # K50: Valor da Fatura (R$)
    ws['K50'] = 'Valor R$'
    # J52: Código do Cliente (Rodapé)
    ws['J52'] = 'Código Cliente (Rodapé)'
    # L52: Número da Fatura (Rodapé)
    ws['L52'] = 'Fatura N° (Rodapé)'

    # Aplica alinhamento e formatação mínima para visualizar
    for cell in ['H9', 'K9', 'A13', 'A15', 'G15', 'B28', 'K50', 'J52', 'L52']:
        ws[cell].alignment = Alignment(wrapText=True)
        ws[cell].font = Font(bold=True)

    # A15 e B28 são campos grandes, vamos garantir que ele tenha um bom tamanho
    ws.column_dimensions[get_column_letter(1)].width = 25
    ws.column_dimensions[get_column_letter(2)].width = 35

    wb.save(MODELO_FILE)
    print(f"'{MODELO_FILE}' criado com sucesso.")

# Garante que a pasta de saída e o modelo existam
os.makedirs(SAIDA_FOLDER, exist_ok=True)
create_initial_xlsx_template()


# --- Classe Principal da Aplicação GUI ---

class CreditNoteApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # --- Setup da Janela Principal ---
        self.title("Sistema de Gestão de Notas de Crédito")
        self.geometry("1000x700")
        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")

        # --- Dados do Backend ---
        self.clientes = load_clientes()
        self.estado = load_estado()
        self.selected_client = None

        # --- Configuração de Layout (Grid) ---
        self.grid_columnconfigure((0, 1), weight=1)
        self.grid_rowconfigure(0, weight=1)

        # --- Frames ---
        self.client_frame = ctk.CTkFrame(self)
        self.client_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        self.note_frame = ctk.CTkFrame(self)
        self.note_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        # Configurações internas dos frames
        self._setup_client_management(self.client_frame)
        self._setup_note_generation(self.note_frame)

    # --- Setup: Gerenciamento de Clientes ---

    def _setup_client_management(self, master):
        master.grid_columnconfigure(0, weight=1)
        master.grid_rowconfigure(5, weight=1) # Row for the listbox

        ctk.CTkLabel(master, text="Gestão de Clientes", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, padx=10, pady=(10, 5), sticky="ew")

        # Campos de busca e ação
        self.client_search_entry = ctk.CTkEntry(master, placeholder_text="Buscar por Código ou Nome")
        self.client_search_entry.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="ew")
        self.client_search_entry.bind("<KeyRelease>", self._filter_client_list)

        button_frame = ctk.CTkFrame(master)
        button_frame.grid(row=2, column=0, padx=10, pady=5, sticky="ew")
        button_frame.grid_columnconfigure((0, 1, 2), weight=1)

        ctk.CTkButton(button_frame, text="Cadastrar", command=self._add_client_dialog).grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        ctk.CTkButton(button_frame, text="Editar", command=self._edit_client_dialog).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ctk.CTkButton(button_frame, text="Excluir", command=self._delete_client).grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        ctk.CTkLabel(master, text="Clientes Encontrados (Clique para Selecionar):", anchor="w").grid(row=3, column=0, padx=10, pady=(5, 0), sticky="ew")

        # Listbox para clientes
        list_frame = ctk.CTkFrame(master)
        list_frame.grid(row=4, column=0, padx=10, pady=(0, 10), sticky="nsew")
        list_frame.grid_columnconfigure(0, weight=1)
        list_frame.grid_rowconfigure(0, weight=1)

        # Usamos Listbox do Tkinter puro para melhor integração com Scrollbar padrão
        # Configurando foreground e background para se misturar melhor com CTk
        self.client_listbox = Listbox(
            list_frame, 
            height=20, 
            borderwidth=0, 
            highlightthickness=0, 
            selectmode=tk.SINGLE, 
            font=("Inter", 12),
            fg="#202020", # Cor de texto escura
            selectbackground="#0074D9", # Azul de seleção
            selectforeground="white"
        )
        self.client_listbox.grid(row=0, column=0, sticky="nsew")

        scrollbar = Scrollbar(list_frame, orient="vertical", command=self.client_listbox.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.client_listbox.config(yscrollcommand=scrollbar.set)

        self.client_listbox.bind('<<ListboxSelect>>', self._select_client_from_list)
        self._update_client_list()


    def _update_client_list(self, filter_text=""):
        """
        Atualiza a Listbox de clientes, ordenando por código e aplicando cor zebrada.
        """
        self.client_listbox.delete(0, tk.END)
        self.filtered_clients = []
        filter_text = filter_text.lower()
        
        # 1. Ordena a lista de clientes por 'codigo' antes de filtrar
        sorted_clients = sorted(self.clientes, key=lambda c: c['codigo'].lower())
        
        # Variável para rastrear o índice dos itens filtrados (para aplicar cor zebrada)
        listbox_index = 0

        for client in sorted_clients:
            if filter_text in client['codigo'].lower() or filter_text in client['nome'].lower():
                display_text = f"[{client['codigo']}] - {client['nome']}"
                
                # 2. Insere o item
                self.client_listbox.insert(tk.END, display_text)
                
                # 3. Aplica a cor zebrada (alternando entre branco e cinza claro)
                if listbox_index % 2 == 0:
                    bg_color = COLOR_DARK
                else:
                    bg_color = COLOR_LIGHT
                
                self.client_listbox.itemconfig(tk.END, {'bg': bg_color})
                
                self.filtered_clients.append(client)
                listbox_index += 1 # Incrementa apenas se o item foi adicionado à Listbox

    def _filter_client_list(self, event):
        """Disparado pela digitação na caixa de busca."""
        self._update_client_list(self.client_search_entry.get())

    def _select_client_from_list(self, event):
        """Seleciona um cliente da Listbox e atualiza o frame da nota."""
        try:
            selection = self.client_listbox.curselection()
            if not selection:
                return

            index = selection[0]
            self.selected_client = self.filtered_clients[index]
            self.client_code_var.set(self.selected_client['codigo'])
            self.client_name_label.configure(text=self.selected_client['nome'])
            
            messagebox.showinfo("Cliente Selecionado", f"Cliente {self.selected_client['nome']} selecionado para a nota.")

        except Exception as e:
            print(f"Erro ao selecionar cliente: {e}")
            messagebox.showerror("Erro", "Falha ao selecionar o cliente.")
            self.selected_client = None

    # Gerenciamento de Modal (Cadastrar/Editar)

    def _add_client_dialog(self):
        """Exibe um diálogo para cadastrar um novo cliente."""
        self._show_client_modal("Cadastrar Cliente", None)

    def _edit_client_dialog(self):
        """Exibe um diálogo para editar o cliente selecionado."""
        if not self.selected_client:
            messagebox.showwarning("Atenção", "Selecione um cliente na lista para editar.")
            return

        self._show_client_modal("Editar Cliente", self.selected_client)

    def _show_client_modal(self, title, client_data):
        """Lógica comum do modal de Cadastro/Edição de Cliente."""
        modal = Toplevel(self)
        modal.title(title)
        modal.transient(self) # Foca na janela principal
        modal.grab_set()
        modal.resizable(False, False)
        modal.geometry("350x250") # Aumentado para acomodar os campos

        # Centralizar o modal
        self.update_idletasks()
        x = self.winfo_x() + self.winfo_width() // 2 - modal.winfo_width() // 2
        y = self.winfo_y() + self.winfo_height() // 2 - modal.winfo_height() // 2
        modal.geometry(f'+{x}+{y}')

        ctk.CTkLabel(modal, text="Código Único:").pack(pady=(10, 0), padx=10)
        code_entry = ctk.CTkEntry(modal, width=300)
        code_entry.pack(padx=10)

        ctk.CTkLabel(modal, text="Nome (Razão Social):").pack(pady=(10, 0), padx=10)
        name_entry = ctk.CTkEntry(modal, width=300)
        name_entry.pack(padx=10)

        original_code = None
        if client_data:
            original_code = client_data['codigo']
            code_entry.insert(0, original_code)
            name_entry.insert(0, client_data['nome'])

        def save_client():
            new_code = code_entry.get().strip()
            name = name_entry.get().strip()

            if not new_code or not name:
                messagebox.showerror("Erro", "Código e Nome são obrigatórios.")
                return

            if client_data:
                # Lógica de Edição
                
                # 1. Verifica se o novo código é diferente do original E se já existe na lista (exceto o próprio cliente)
                is_duplicate = any(c['codigo'] == new_code and c['codigo'] != original_code for c in self.clientes)
                
                if is_duplicate:
                    messagebox.showerror("Erro", f"O Código de Cliente '{new_code}' já existe para outro cliente.")
                    return
                
                # 2. Atualiza os dados
                client_data['codigo'] = new_code
                client_data['nome'] = name
                messagebox.showinfo("Sucesso", f"Cliente {new_code} atualizado.")
                
                # Se o cliente selecionado for o que está sendo editado, atualiza a UI principal
                if self.selected_client and original_code == self.selected_client['codigo']:
                    self.client_code_var.set(new_code)
                    self.client_name_label.configure(text=name)

            else:
                # Lógica de Cadastro
                if any(c['codigo'] == new_code for c in self.clientes):
                    messagebox.showerror("Erro", f"O Código de Cliente '{new_code}' já existe.")
                    return
                self.clientes.append({"codigo": new_code, "nome": name})
                messagebox.showinfo("Sucesso", f"Cliente {new_code} cadastrado.")

            save_clientes(self.clientes)
            # Atualiza a lista com o novo cliente ou código ordenado
            self._update_client_list(self.client_search_entry.get())
            modal.destroy()

        ctk.CTkButton(modal, text="Salvar", command=save_client).pack(pady=20, padx=10)

    def _delete_client(self):
        """Exclui o cliente selecionado."""
        if not self.selected_client:
            messagebox.showwarning("Atenção", "Selecione um cliente para excluir.")
            return

        code = self.selected_client['codigo']
        name = self.selected_client['nome']

        if messagebox.askyesno("Confirmar Exclusão", f"Tem certeza que deseja excluir o cliente:\n[{code}] - {name}?"):
            self.clientes = [c for c in self.clientes if c['codigo'] != code]
            save_clientes(self.clientes)
            self.selected_client = None
            self.client_code_var.set("")
            self.client_name_label.configure(text="Nenhum cliente selecionado")
            self._update_client_list(self.client_search_entry.get())
            messagebox.showinfo("Sucesso", f"Cliente {code} excluído com sucesso.")

    # --- Setup: Geração de Nota de Crédito ---

    def _setup_note_generation(self, master):
        master.grid_columnconfigure(0, weight=1)
        master.grid_rowconfigure(5, weight=1) # Row for the description Textbox

        ctk.CTkLabel(master, text="Geração de Nota de Crédito", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, padx=10, pady=(10, 5), sticky="ew")

        # 1. Dados da Fatura (Data e Número)
        data_frame = ctk.CTkFrame(master)
        data_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        data_frame.grid_columnconfigure((0, 1), weight=1)

        # Data
        ctk.CTkLabel(data_frame, text="Data (DD/MM/AAAA):").grid(row=0, column=0, padx=10, pady=(5, 0), sticky="w")
        self.date_var = tk.StringVar(value=datetime.date.today().strftime("%d/%m/%Y"))
        self.date_var.trace_add("write", self._format_date_input)
        self.date_entry = ctk.CTkEntry(data_frame, textvariable=self.date_var)
        self.date_entry.grid(row=1, column=0, padx=10, pady=(0, 5), sticky="ew")

        # Número da Fatura
        ctk.CTkLabel(data_frame, text="Número da Fatura (Próx. Sugerido: {}):"
            .format(self.estado['ultima_fatura'])).grid(row=0, column=1, padx=10, pady=(5, 0), sticky="w")
        self.invoice_number_var = ctk.StringVar(value=str(self.estado['ultima_fatura']))
        self.invoice_number_entry = ctk.CTkEntry(data_frame, textvariable=self.invoice_number_var)
        self.invoice_number_entry.grid(row=1, column=1, padx=10, pady=(0, 5), sticky="ew")

        # 2. Dados do Cliente
        client_info_frame = ctk.CTkFrame(master)
        client_info_frame.grid(row=2, column=0, padx=10, pady=5, sticky="ew")
        client_info_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(client_info_frame, text="Cliente Selecionado:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=(5, 0), sticky="w")

        self.client_code_var = ctk.StringVar(value="")
        ctk.CTkLabel(client_info_frame, text="Código:").grid(row=1, column=0, padx=10, pady=(0, 0), sticky="w")
        ctk.CTkEntry(client_info_frame, textvariable=self.client_code_var, state="readonly").grid(row=2, column=0, padx=10, pady=(0, 5), sticky="ew")

        ctk.CTkLabel(client_info_frame, text="Nome (Razão Social):").grid(row=3, column=0, padx=10, pady=(0, 0), sticky="w")
        self.client_name_label = ctk.CTkLabel(client_info_frame, text="Nenhum cliente selecionado", anchor="w", fg_color=("gray85", "gray25"), corner_radius=5)
        self.client_name_label.grid(row=4, column=0, padx=10, pady=(0, 10), sticky="ew", ipady=5)


        # 3. Valor e Descrição
        value_desc_frame = ctk.CTkFrame(master)
        value_desc_frame.grid(row=3, column=0, padx=10, pady=5, sticky="ew")
        value_desc_frame.grid_columnconfigure(0, weight=1)

        # Valor da Fatura
        ctk.CTkLabel(value_desc_frame, text="Valor da Fatura (R$):").grid(row=0, column=0, padx=10, pady=(5, 0), sticky="w")
        self.value_var = tk.StringVar(value="0,00")
        self.value_var.trace_add("write", self._format_currency_input)
        self.value_entry = ctk.CTkEntry(value_desc_frame, textvariable=self.value_var)
        self.value_entry.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="ew")

        # Descrição/Histórico
        ctk.CTkLabel(master, text=f"Descrição/Histórico (Última nota: '{self.estado['ultima_descricao'][:30]}...')").grid(row=4, column=0, padx=10, pady=(5, 0), sticky="w")
        
        self.description_textbox = ctk.CTkTextbox(master, height=150)
        self.description_textbox.grid(row=5, column=0, padx=10, pady=(0, 10), sticky="nsew")
        self.description_textbox.insert("1.0", self.estado['ultima_descricao'])

        # 4. Botão de Ação
        action_frame = ctk.CTkFrame(master)
        action_frame.grid(row=6, column=0, padx=10, pady=10, sticky="ew")
        action_frame.grid_columnconfigure((0, 1), weight=1)

        ctk.CTkButton(action_frame, text="Gerar Nota de Crédito (.xlsx)", command=self._process_note, fg_color="green").grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        ctk.CTkButton(action_frame, text="Imprimir Última Nota", command=self._print_last_note, fg_color="gray").grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        self.last_saved_file = None # Para rastrear o arquivo a ser impresso

    # --- Funções de Formatação e Validação de Input ---

    def _format_date_input(self, *args):
        """
        Formata o campo de data em DD/MM/AAAA em tempo real (Versão corrigida e robusta).
        """
        current = self.date_var.get()
        cursor_pos = self.date_entry.index(tk.INSERT)
        
        # 1. Limpa o valor para obter apenas dígitos, limitando a 8 (DDMMYYYY)
        numeric = re.sub(r'[^0-9]', '', current)
        numeric = numeric[:8]

        # 2. Reconstroi o valor formatado inserindo as barras nas posições corretas
        new_value = ""
        if len(numeric) > 0:
            new_value += numeric[0:2]
        if len(numeric) > 2:
            new_value += "/" + numeric[2:4]
        if len(numeric) > 4:
            new_value += "/" + numeric[4:8]
        
        # 3. Calcula a nova posição do cursor com base na posição numérica original
        
        # Quantidade de dígitos no texto original antes da posição do cursor
        old_numeric_prefix = re.sub(r'[^0-9]', '', current[:cursor_pos])
        target_numeric_length = len(old_numeric_prefix)
        
        new_cursor_pos = 0
        numeric_count = 0
        
        # Itera sobre a nova string para encontrar onde o dígito final da parte antiga parou
        for char in new_value:
            if char.isdigit():
                numeric_count += 1
            
            new_cursor_pos += 1
            
            if numeric_count == target_numeric_length:
                break
        
        # Casos especiais de ajuste do cursor:
        if target_numeric_length == 0:
            new_cursor_pos = 0
        elif cursor_pos >= len(current):
            # Se o cursor estava no final, ele deve permanecer no final
            new_cursor_pos = len(new_value)
        elif len(new_value) > len(current) and new_value[new_cursor_pos - 1] == '/':
             # Se uma barra foi inserida e o cursor está logo após o dígito, avança sobre a barra
             new_cursor_pos += 1

        # 4. Atualiza o valor e a posição do cursor se o valor mudou
        if new_value != current:
            self.date_var.set(new_value)
            try:
                self.date_entry.icursor(new_cursor_pos)
            except:
                pass 
        else:
            # Mantém a posição original se o valor não mudou (e.g., apagando)
            try:
                self.date_entry.icursor(cursor_pos)
            except:
                pass


    def _format_currency_input(self, *args):
        """Formata o campo de valor em R$ X.XXX,XX em tempo real."""
        current = self.value_var.get()
        cursor_pos = self.value_entry.index(tk.INSERT)

        # 1. Limpa o valor para apenas dígitos
        cleaned_value = re.sub(r'[^0-9]', '', current)

        if not cleaned_value:
            new_value = "0,00"
            new_cursor_pos = 3 # Posiciona após '0,00'
        else:
            # 2. Converte para centavos e formata
            value_cents = int(cleaned_value)
            value_reais = value_cents / 100

            # Formata o valor com pontos de milhar e vírgula decimal
            new_value = f"{value_reais:,.2f}".replace('.', '#').replace(',', '.').replace('#', ',')
            
            # Calculando a nova posição do cursor (tentativa simples)
            # Remove a contagem de pontos/vírgulas do valor original para achar a posição numérica
            original_numeric_len = len(re.sub(r'[^0-9]', '', current[:cursor_pos]))
            
            new_cursor_pos = 0
            numeric_count = 0
            for i, char in enumerate(new_value):
                if char.isdigit():
                    numeric_count += 1
                    if numeric_count > original_numeric_len:
                        break
                new_cursor_pos = i + 1
            
            # Se o cursor estava no final, mantém no final
            if cursor_pos >= len(current):
                new_cursor_pos = len(new_value)


        # 3. Atualiza o valor e tenta manter a posição do cursor
        if new_value != current:
            self.value_var.set(new_value)
            try:
                self.value_entry.icursor(new_cursor_pos)
            except:
                pass # Ignora se o cursor não puder ser definido

    # --- Funções de Processamento da Nota ---

    def _process_note(self):
        """Valida, preenche o XLSX, salva e atualiza o estado."""
        # 1. Validação dos Dados
        if not self.selected_client:
            messagebox.showerror("Erro de Validação", "Selecione um cliente para prosseguir.")
            return

        data_input = self.date_var.get()
        invoice_number = self.invoice_number_var.get()
        description_text = self.description_textbox.get("1.0", tk.END).strip()
        value_formatted = self.value_var.get()

        if not re.fullmatch(r'\d{2}/\d{2}/\d{4}', data_input):
            messagebox.showerror("Erro de Validação", "Formato de Data inválido. Use DD/MM/AAAA.")
            return
        
        if not invoice_number.isdigit() or int(invoice_number) <= 0:
            messagebox.showerror("Erro de Validação", "Número da Fatura deve ser um número inteiro positivo.")
            return

        # Limpa o valor formatado para um float
        try:
            value_cleaned = value_formatted.replace('.', '').replace(',', '.')
            value_float = float(value_cleaned)
        except ValueError:
            messagebox.showerror("Erro de Validação", "Valor da Fatura inválido.")
            return

        if not description_text:
            messagebox.showerror("Erro de Validação", "A Descrição/Histórico é obrigatória.")
            return
        
        # 2. Carregar e Preencher o Modelo XLSX
        try:
            wb = load_workbook(MODELO_FILE)
            ws = wb.active
        except FileNotFoundError:
            messagebox.showerror("Erro de Arquivo", f"O arquivo modelo '{MODELO_FILE}' não foi encontrado.")
            return
        except Exception as e:
            messagebox.showerror("Erro de Leitura", f"Erro ao abrir o arquivo modelo: {e}")
            return

        # Mapeamento e Preenchimento das Células
        client_code = self.selected_client['codigo']
        client_name = self.selected_client['nome']

        # --- GERAÇÃO DO NOME DA PLANILHA (ABA) ---
        # 1. Limpa o nome para pegar apenas palavras
        name_parts = re.sub(r'[^\w\s]', '', client_name).split()
        
        # 2. Pega as duas primeiras palavras (ou menos, se não houver)
        sheet_name_base = '_'.join(name_parts[:2]).upper()
        
        # 3. Combina e garante que não excede o limite do Excel (31 caracteres)
        new_sheet_title = f"{sheet_name_base}_{invoice_number}"
        
        # Limita a 31 caracteres (máximo do Excel)
        if len(new_sheet_title) > 31:
            new_sheet_title = new_sheet_title[:31]

        # 4. Renomeia a planilha ativa
        try:
            ws.title = new_sheet_title
        except Exception as e:
            # Caso o nome seja inválido por algum motivo (caractere não permitido, etc.)
            print(f"Aviso: Não foi possível renomear a planilha para '{new_sheet_title}': {e}. Usando o nome original.")


        # Campos e Células
        cell_map = {
            # Data (DD/MM/AAAA)
            'H9': data_input,
            # Número da Fatura (Topo e Rodapé)
            'K9': invoice_number,
            'L52': invoice_number,
            # Código do Cliente (Topo e Rodapé)
            'A13': client_code,
            'J52': client_code,
            # Nome do Cliente (Razão Social - Faturar a:)
            'A15': client_name,
            # Nome do Cliente (Razão Social - Enviado a/Consignatário: )
            'G15': client_name, 
            # Descrição/Histórico
            'B28': description_text,
            # Valor da Fatura (R$) - Formato com duas casas decimais
            'K50': value_float # openpyxl aceita float, formataremos com number_format
        }

        # Aplica os valores na planilha
        try:
            for cell, value in cell_map.items():
                ws[cell] = value
                # Se for o valor, define o formato de moeda para o Excel
                if cell == 'K50':
                    ws[cell].number_format = 'R$ #,##0.00'
        except Exception as e:
            messagebox.showerror("Erro de Preenchimento", f"Falha ao preencher a célula. Verifique se o '{MODELO_FILE}' está fechado e acessível.\nDetalhe: {e}")
            return

        # 3. Salvar a Nova Nota
        file_name = f"Nota_Credito_{invoice_number}_Cliente_{client_code}.xlsx"
        output_path = os.path.join(SAIDA_FOLDER, file_name)

        try:
            wb.save(output_path)
            self.last_saved_file = output_path
        except Exception as e:
            messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar a nota em '{output_path}'. Verifique as permissões.\nDetalhe: {e}")
            return

        # 4. Atualizar o Estado do Programa
        try:
            # Incrementa a fatura e salva a nova descrição
            next_invoice = int(invoice_number) + 1
            self.estado['ultima_fatura'] = next_invoice
            self.estado['ultima_descricao'] = description_text
            save_estado(self.estado)

            # Atualiza a GUI com o novo sugerido
            self.invoice_number_var.set(str(next_invoice))
            # O label precisa ser atualizado, pegamos o container (data_frame) e atualizamos o label [0]
            # O índice 2 é o label do número da fatura dentro do data_frame
            data_frame_children = self.note_frame.winfo_children()[0].winfo_children()
            if len(data_frame_children) > 2:
                data_frame_children[2].configure(text="Número da Fatura (Próx. Sugerido: {}):".format(self.estado['ultima_fatura']))
            
            # Limpa e preenche a descrição para a próxima nota
            self.description_textbox.delete("1.0", tk.END)
            self.description_textbox.insert("1.0", self.estado['ultima_descricao'])
            self.value_var.set("0,00")
            
        except Exception as e:
            print(f"Aviso: Falha ao atualizar o estado do programa: {e}")
            # Não é um erro crítico que impeça o salvamento da nota.

        # 5. Sucesso e Impressão (Abre diretamente)
        messagebox.showinfo("Nota Gerada com Sucesso", f"Nota salva em:\n{output_path}\n\nO comando de impressão será enviado agora.")
        self._print_file(output_path)
            
    def _print_file(self, filepath):
        """Tenta abrir o arquivo com o programa padrão (o que geralmente abre a caixa de diálogo de impressão)."""
        try:
            if os.name == 'nt':  # Windows
                # Usa startfile para abrir o arquivo com o aplicativo padrão
                # Em muitos sistemas, isso abrirá a caixa de diálogo de impressão se for um arquivo de documento
                os.startfile(filepath, "print")
            elif os.uname()[0] == 'Darwin':  # macOS
                # Usa 'lp' para enviar para a impressora padrão (ou 'open -p' para abrir o diálogo)
                os.system(f'open -a "Microsoft Excel" -p "{filepath}"')
            else:  # Linux (genérico)
                # Tenta abrir o arquivo
                os.system(f'xdg-open "{filepath}"')
                messagebox.showinfo("Impressão", "Comando de impressão enviado. Verifique a fila de impressão do sistema.")
        except Exception as e:
            messagebox.showerror("Erro de Impressão/Abertura", f"Não foi possível abrir o arquivo para impressão. Tente abrir o arquivo manualmente: {filepath}\nDetalhe: {e}")

    def _print_last_note(self):
        """Imprime o último arquivo salvo."""
        if not self.last_saved_file or not os.path.exists(self.last_saved_file):
            messagebox.showwarning("Atenção", "Nenhuma nota foi gerada nesta sessão ou o último arquivo salvo não foi encontrado.")
            return

        if messagebox.askyesno("Confirmar Impressão", f"Deseja enviar para impressão o arquivo:\n{os.path.basename(self.last_saved_file)}?"):
            self._print_file(self.last_saved_file)


if __name__ == "__main__":
    app = CreditNoteApp()
    app.mainloop()
